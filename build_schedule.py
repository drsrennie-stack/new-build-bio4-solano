#!/usr/bin/env python3
"""
Course Schedule Engine, build script (schema 3.0).

One source of truth, many outputs. Edit schedule-data.xlsx (or schedule-data.json),
then run this script. It rebuilds:

  course-schedule.html        the interactive page, with fresh data baked in
  schedule-data.json          the canonical data file
  course-schedule.pdf         a print-ready schedule
  exports/*.ics               calendar file, due dates
  exports/*-spaced-repetition.csv   starter Anki deck, one card per lab topic
  exports/*-study-plan.txt    daily prep plus scheduled spaced reviews
  exports/*-planner-checklist.md    Markdown checklist for any task planner
  exports/*-slo-grade-map.csv every graded item mapped to an SLO

Day model (schema 3.0):
  Each day has a Lecture track, a Lab track, and a nightly prework line.
  Each week has a weekend card (Friday to Sunday prework). Graded work is the
  lab exams and the TBLs. Spaced recall daily review starts the day after the
  first class day, since no cards exist yet on day one.

Usage:
  python build_schedule.py                 read the xlsx if present, else the json
  python build_schedule.py --source json   force the json as the source
  python build_schedule.py --source xlsx   force the xlsx as the source
  python build_schedule.py --no-pdf        skip the PDF (if reportlab is missing)

No network access, no student data, no external services.
Prepared for Dr. Sharilyn Rennie.
"""
import os
import re
import sys
import csv
import json
import io
import datetime as dt

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(BASE, "schedule-data.xlsx")
JSON = os.path.join(BASE, "schedule-data.json")
HTML = os.path.join(BASE, "course-schedule.html")
PDF = os.path.join(BASE, "course-schedule.pdf")
EXPORT_DIR = os.path.join(BASE, "exports")

ITEM_TYPES = {"lab-exam": "Lab Practical Exam", "tbl": "Team-Based Learning"}
EXAM_TYPES = {"lab-exam"}
SR_NOTE = "Review offsets, in days, after a topic is first taught."
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTHS_LONG = ["January", "February", "March", "April", "May", "June", "July",
               "August", "September", "October", "November", "December"]
WEEKDAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
NAVY_HEX, TERRA_HEX = "#1E3D4C", "#A0522D"

WARNINGS = []


def warn(msg):
    WARNINGS.append(msg)


# ----------------------------------------------------------------------
# Date helpers
# ----------------------------------------------------------------------
def parse_iso(s):
    return dt.date(*[int(x) for x in s.split("-")])


def norm_date(v):
    if v is None:
        return ""
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip().replace("/", "-")


def fmt_short(iso):
    d = parse_iso(iso)
    return WEEKDAYS[d.weekday()][:3] + ", " + MONTHS[d.month - 1] + " " + str(d.day)


def fmt_long(iso):
    d = parse_iso(iso)
    return MONTHS_LONG[d.month - 1] + " " + str(d.day) + ", " + str(d.year)


def fmt_range(a, b):
    da, db = parse_iso(a), parse_iso(b)
    if da.month == db.month:
        return "%s %d to %d" % (MONTHS[da.month - 1], da.day, db.day)
    return "%s %d to %s %d" % (MONTHS[da.month - 1], da.day, MONTHS[db.month - 1], db.day)


def add_days(iso, n):
    return (parse_iso(iso) + dt.timedelta(days=n)).strftime("%Y-%m-%d")


def weekday_name(iso):
    return WEEKDAYS[parse_iso(iso).weekday()]


def has_new_cards(day):
    """New spaced-recall cards seed on class days with non-exam lab content."""
    if not day.get("lab"):
        return False
    return not any(i["type"] in EXAM_TYPES for i in day.get("items", []))


# ----------------------------------------------------------------------
# Source readers
# ----------------------------------------------------------------------
def read_json(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _items(cells):
    """Build a list of {label,url} from alternating label/url cells."""
    out = []
    for i in range(0, len(cells) - 1, 2):
        label, url = cells[i], cells[i + 1]
        if label not in (None, ""):
            entry = {"label": str(label).strip()}
            if url not in (None, ""):
                entry["url"] = str(url).strip()
            out.append(entry)
    return out


def read_xlsx(path):
    """Reconstruct the schedule data dict from the spreadsheet template."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)

    # ---- Course sheet ----
    cs = wb["Course"]
    fields = {}
    for row in cs.iter_rows(min_row=3, max_col=2, values_only=True):
        if row[0]:
            fields[str(row[0]).strip()] = row[1]
    sample = str(fields.get("isSampleData", "FALSE")).strip().upper() in ("TRUE", "1", "YES")
    intervals = []
    for part in str(fields.get("srIntervals", "1, 3, 8, 21")).split(","):
        part = part.strip()
        if part:
            intervals.append(int(float(part)))

    course = {
        "code": str(fields.get("code", "")).strip(),
        "title": str(fields.get("title", "")).strip(),
        "institution": str(fields.get("institution", "")).strip(),
        "term": str(fields.get("term", "")).strip(),
        "instructor": str(fields.get("instructor", "")).strip(),
        "meetingPattern": str(fields.get("meetingPattern", "")).strip(),
        "format": str(fields.get("format", "")).strip(),
        "timezone": str(fields.get("timezone", "America/Los_Angeles")).strip(),
        "isSampleData": sample,
        "spacedRecallUrl": str(fields.get("spacedRecallUrl", "")).strip(),
        "slos": [],
        "itemTypes": dict(ITEM_TYPES),
    }
    for row in wb["SLOs"].iter_rows(min_row=3, max_col=2, values_only=True):
        if row[0]:
            course["slos"].append({"code": str(row[0]).strip(),
                                   "description": str(row[1] or "").strip()})

    # ---- Schedule sheet ----
    # Week, Theme, Week Start, Week End, Date, Weekday, Session Type,
    # Lecture 1, Lecture 1 URL, Lecture 2, Lecture 2 URL,
    # Lab 1, Lab 1 URL, Lab 2, Lab 2 URL, Lab 3, Lab 3 URL, Nightly Prework
    days_by_date = {}
    week_meta = {}
    for row in wb["Schedule"].iter_rows(min_row=3, values_only=True):
        if row[0] in (None, ""):
            continue
        week = int(row[0])
        date = norm_date(row[4])
        if not date:
            continue
        week_meta.setdefault(week, {"theme": str(row[1] or "").strip(),
                                    "start": norm_date(row[2]), "end": norm_date(row[3])})
        days_by_date[date] = {
            "week": week, "date": date,
            "weekday": str(row[5] or "").strip() or weekday_name(date),
            "sessionType": str(row[6] or "").strip(),
            "lecture": _items(list(row[7:11])),
            "lab": _items(list(row[11:17])),
            "prework": str(row[17] or "").strip() if len(row) > 17 else "",
            "items": [],
        }

    # ---- Graded Items sheet ----
    for row in wb["Graded Items"].iter_rows(min_row=3, values_only=True):
        if row[0] in (None, ""):
            continue
        date = norm_date(row[1])
        item = {
            "id": str(row[0]).strip(),
            "title": str(row[2] or "").strip(),
            "type": str(row[3] or "").strip(),
            "points": int(float(row[4] or 0)),
            "due": norm_date(row[5]) or date,
            "covers": str(row[6] or "").strip(),
            "slo": [s.strip() for s in re.split(r"[;,]", str(row[7] or "")) if s.strip()],
            "note": str(row[8] or "").strip(),
        }
        if date in days_by_date:
            days_by_date[date]["items"].append(item)
        else:
            warn("Graded item '%s' has Date '%s' which matches no Schedule row." % (item["id"], date))

    # ---- Weekend sheet ----
    weekends = {}
    if "Weekend" in wb.sheetnames:
        for row in wb["Weekend"].iter_rows(min_row=3, values_only=True):
            if row[0] in (None, ""):
                continue
            week = int(row[0])
            pw = [str(c).strip() for c in row[3:5] if c not in (None, "")]
            weekends[week] = {"start": norm_date(row[1]), "end": norm_date(row[2]),
                              "prework": pw, "sunday": str(row[5] or "").strip()}

    # ---- Assemble ----
    by_week = {}
    for d in days_by_date.values():
        by_week.setdefault(d["week"], []).append(d)
    weeks = []
    for week in sorted(week_meta):
        meta = week_meta[week]
        wdays = sorted(by_week.get(week, []), key=lambda d: d["date"])
        for d in wdays:
            d.pop("week", None)
        if not meta["start"] and wdays:
            meta["start"] = wdays[0]["date"]
        if not meta["end"] and wdays:
            meta["end"] = wdays[-1]["date"]
        wk = {"week": week, "theme": meta["theme"],
              "start": meta["start"], "end": meta["end"], "days": wdays}
        if week in weekends:
            wk["weekend"] = weekends[week]
        weeks.append(wk)

    return {
        "schemaVersion": "3.0",
        "course": course,
        "weeks": weeks,
        "spacedRepetition": {"intervalsDays": intervals or [1, 3, 8, 21], "note": SR_NOTE},
    }


# ----------------------------------------------------------------------
# Validation
# ----------------------------------------------------------------------
def normalise(data):
    course = data["course"]
    course.setdefault("itemTypes", dict(ITEM_TYPES))
    data.setdefault("spacedRepetition", {"intervalsDays": [1, 3, 8, 21], "note": SR_NOTE})

    slo_codes = {s["code"] for s in course.get("slos", [])}
    all_items = []
    for w in data["weeks"]:
        for day in w["days"]:
            day.setdefault("lecture", [])
            day.setdefault("lab", [])
            day.setdefault("items", [])
            day.setdefault("prework", "")
            day.setdefault("sessionType", "")
            if not day.get("weekday") and day.get("date"):
                day["weekday"] = weekday_name(day["date"])
            for it in day["items"]:
                all_items.append(it)
                if it["type"] not in ITEM_TYPES:
                    warn("Item '%s' has type '%s'. Expected lab-exam or tbl."
                         % (it["id"], it["type"]))
                for code in it.get("slo", []):
                    if slo_codes and code not in slo_codes:
                        warn("Item '%s' references SLO '%s' not defined on the SLOs tab."
                             % (it["id"], code))
                try:
                    parse_iso(it["due"])
                except Exception:
                    warn("Item '%s' has an unreadable due date: '%s'." % (it["id"], it.get("due")))
    course["totalPoints"] = sum(int(i.get("points", 0)) for i in all_items)
    return data


# ----------------------------------------------------------------------
# Writers
# ----------------------------------------------------------------------
def write_json(data, path):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
        f.write("\n")


def update_html(data, path):
    if not os.path.exists(path):
        warn("course-schedule.html not found, skipped the HTML refresh.")
        return False
    with open(path, encoding="utf-8") as f:
        html = f.read()
    payload = json.dumps(data, indent=2, ensure_ascii=False)
    pattern = re.compile(
        r'(<script id="schedule-data" type="application/json">)(.*?)(</script>)', re.DOTALL)
    if not pattern.search(html):
        warn("Could not find the schedule-data block in the HTML, skipped.")
        return False
    with open(path, "w", encoding="utf-8") as f:
        f.write(pattern.sub(lambda m: m.group(1) + "\n" + payload + "\n" + m.group(3), html))
    return True


# ----------------------------------------------------------------------
# Export builders
# ----------------------------------------------------------------------
def iter_items(data):
    rows = []
    for w in data["weeks"]:
        for day in w["days"]:
            for it in day["items"]:
                rows.append({"item": it, "date": day["date"], "week": w["week"]})
    rows.sort(key=lambda r: r["date"])
    return rows


def csv_string(rows):
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\r\n")
    for r in rows:
        w.writerow(r)
    return buf.getvalue()


def ics_escape(s):
    return (str(s).replace("\\", "\\\\").replace(";", "\\;")
            .replace(",", "\\,").replace("\n", "\\n"))


def ics_fold(line):
    if len(line) <= 73:
        return line
    out, i = "", 0
    while i < len(line):
        size = 73 if i == 0 else 72
        out += ("" if i == 0 else "\r\n ") + line[i:i + size]
        i += size
    return out


def build_ics(data):
    course = data["course"]
    stamp = dt.datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    L = ["BEGIN:VCALENDAR", "VERSION:2.0",
         "PRODID:-//MedMasters Collaborative//Course Schedule Engine//EN",
         "CALSCALE:GREGORIAN", "METHOD:PUBLISH",
         "X-WR-CALNAME:" + ics_escape("%s %s due dates" % (course["code"], course["title"]))]
    for r in iter_items(data):
        it = r["item"]
        d = it["due"].replace("-", "")
        d_end = add_days(it["due"], 1).replace("-", "")
        label = ITEM_TYPES.get(it["type"], it["type"])
        desc = "%s. Worth %d points. %s%sSLOs: %s." % (
            label, it["points"],
            ("Covers " + it["covers"] + ". ") if it.get("covers") else "",
            (it["note"] + " ") if it.get("note") else "",
            ", ".join(it.get("slo", [])))
        L += ["BEGIN:VEVENT",
              "UID:%s-%s@course-schedule-engine" % (it["id"], d),
              "DTSTAMP:" + stamp,
              "DTSTART;VALUE=DATE:" + d,
              "DTEND;VALUE=DATE:" + d_end,
              ics_fold("SUMMARY:" + ics_escape("%s: %s (%d pts)" % (course["code"], it["title"], it["points"]))),
              ics_fold("DESCRIPTION:" + ics_escape(desc)),
              "CATEGORIES:" + ics_escape(label),
              "TRANSP:TRANSPARENT"]
        if it["type"] in EXAM_TYPES:
            L += ["BEGIN:VALARM", "ACTION:DISPLAY", "TRIGGER:-P1D",
                  ics_fold("DESCRIPTION:" + ics_escape("Tomorrow: " + it["title"])), "END:VALARM"]
        L.append("END:VEVENT")
    L.append("END:VCALENDAR")
    return "\r\n".join(L) + "\r\n"


def slug(s):
    return re.sub(r"^-|-$", "", re.sub(r"[^a-z0-9]+", "-", s.lower()))


def build_anki(data):
    course = data["course"]
    lines = ["#separator:Comma", "#html:false", "#columns:Front Back Deck Tags"]
    rows = []
    code = course["code"].replace(" ", "")
    for w in data["weeks"]:
        deck = "%s::Week %d %s" % (code, w["week"], w["theme"].replace(":", " "))
        for day in w["days"]:
            if not has_new_cards(day):
                continue
            for l in day["lab"]:
                rows.append(["Identify and describe the structures of: " + l["label"],
                             "From the %s session on %s." % (day.get("sessionType", "class"),
                                                             fmt_short(day["date"])),
                             deck,
                             "BIO004 week%d %s" % (w["week"], slug(l["label"]))])
    return "\n".join(lines) + "\n" + csv_string(rows).replace("\r\n", "\n")


def build_study_plan(data):
    course = data["course"]
    intervals = data["spacedRepetition"]["intervalsDays"]
    course_start = data["weeks"][0]["start"]
    course_end = data["weeks"][-1]["end"]
    review_by_date = {}
    for w in data["weeks"]:
        for day in w["days"]:
            if not has_new_cards(day):
                continue
            for l in day["lab"]:
                for off in intervals:
                    rd = add_days(day["date"], off)
                    if rd > course_end:
                        continue
                    review_by_date.setdefault(rd, []).append(l["label"])

    L = ["%s %s, Daily Prep and Study Plan" % (course["code"], course["title"]),
         "%s. Prepared by Dr. Sharilyn Rennie." % course["term"],
         "Prep work is recommended and not graded. Spaced reviews use offsets of %s days."
         % ", ".join(str(x) for x in intervals),
         "Paste this into any planner or spaced-repetition app.", ""]
    for w in data["weeks"]:
        L += ["==============================",
              "WEEK %d: %s  (%s)" % (w["week"], w["theme"], fmt_range(w["start"], w["end"])),
              "=============================="]
        for day in w["days"]:
            L.append("")
            L.append("%s, %s  ::  %s" % (day["weekday"], fmt_long(day["date"]),
                                         day.get("sessionType", "Class day")))
            for it in day["items"]:
                L.append("  GRADED, due today: %s (%d pts)%s"
                         % (it["title"], it["points"],
                            (", covers " + it["covers"]) if it.get("covers") else ""))
            if day.get("lecture"):
                L.append("  Lecture:")
                for l in day["lecture"]:
                    L.append("    - %s%s" % (l["label"], ("  ->  " + l["url"]) if l.get("url") else ""))
            if day.get("lab"):
                L.append("  Lab:")
                for l in day["lab"]:
                    L.append("    - %s%s" % (l["label"], ("  ->  " + l["url"]) if l.get("url") else ""))
            if day.get("prework"):
                L.append("  Nightly prework (unlocks after the lecture video):")
                L.append("    [ ] " + day["prework"])
            L.append("  Spaced recall:")
            if day["date"] != course_start:
                L.append("    [ ] Daily review, every card due today")
            if has_new_cards(day):
                L.append("    [ ] New cards from today's lab")
            rev = review_by_date.get(day["date"])
            if rev:
                seen = []
                for t in rev:
                    if t not in seen:
                        seen.append(t)
                L.append("  Spaced review of earlier topics:")
                for t in seen:
                    L.append("    [ ] Re-test yourself on: " + t)
        we = w.get("weekend")
        if we:
            L.append("")
            L.append("Weekend  ::  " + fmt_range(we["start"], we["end"]))
            if we.get("prework"):
                L.append("  Weekend prework:")
                for p in we["prework"]:
                    L.append("    [ ] " + p)
            if we.get("sunday"):
                L.append("  Sunday prework:")
                L.append("    [ ] " + we["sunday"])
            L.append("  Spaced recall:")
            L.append("    [ ] Daily review, every card due today")
        L.append("")
    return "\n".join(L) + "\n"


def build_checklist(data):
    course = data["course"]
    course_start = data["weeks"][0]["start"]
    n_items = sum(len(d["items"]) for w in data["weeks"] for d in w["days"])
    L = ["# %s %s, Course Checklist" % (course["code"], course["title"]), "",
         "_%s. %d graded items, %d points. Prepared by Dr. Sharilyn Rennie._"
         % (course["term"], n_items, course["totalPoints"]), ""]
    for w in data["weeks"]:
        L.append("## Week %d: %s  (%s)" % (w["week"], w["theme"], fmt_range(w["start"], w["end"])))
        L.append("")
        for day in w["days"]:
            L.append("### %s, %s" % (fmt_short(day["date"]), day.get("sessionType", "Class day")))
            L.append("")
            for it in day["items"]:
                L.append("- [ ] **DUE %s, %s:** %s (%d pts)%s"
                         % (fmt_short(it["due"]), ITEM_TYPES.get(it["type"], it["type"]),
                            it["title"], it["points"],
                            (", covers " + it["covers"]) if it.get("covers") else ""))
            if day.get("lecture"):
                L.append("- Lecture:")
                for l in day["lecture"]:
                    L.append("  - " + ("[%s](%s)" % (l["label"], l["url"]) if l.get("url") else l["label"]))
            if day.get("lab"):
                L.append("- Lab:")
                for l in day["lab"]:
                    L.append("  - " + ("[%s](%s)" % (l["label"], l["url"]) if l.get("url") else l["label"]))
            if day.get("prework"):
                L.append("- Nightly prework (no link, unlocks after the lecture video):")
                L.append("  - [ ] " + day["prework"])
            L.append("- Spaced recall:")
            if day["date"] != course_start:
                L.append("  - [ ] Daily review, every card due today")
            if has_new_cards(day):
                L.append("  - [ ] New cards from today's lab")
            L.append("")
        we = w.get("weekend")
        if we:
            L.append("### Weekend, %s" % fmt_range(we["start"], we["end"]))
            L.append("")
            if we.get("prework"):
                L.append("- Weekend prework:")
                for p in we["prework"]:
                    L.append("  - [ ] " + p)
            if we.get("sunday"):
                L.append("- Sunday prework:")
                L.append("  - [ ] " + we["sunday"])
            L.append("- Spaced recall:")
            L.append("  - [ ] Daily review, every card due today")
            L.append("")
    return "\n".join(L) + "\n"


def build_slo_csv(data):
    course = data["course"]
    total = course["totalPoints"] or 1
    slo_map = {s["code"]: s["description"] for s in course.get("slos", [])}
    rows = [["Item ID", "Item", "Type", "Week", "Covers", "Due Date", "Points",
             "Percent of Course", "SLO Codes", "SLO Descriptions"]]
    for r in iter_items(data):
        it = r["item"]
        codes = it.get("slo", [])
        rows.append([it["id"], it["title"], ITEM_TYPES.get(it["type"], it["type"]),
                     "Week %d" % r["week"], it.get("covers", ""), it["due"], it["points"],
                     "%.1f%%" % (it["points"] / total * 100),
                     "; ".join(codes),
                     "; ".join(slo_map.get(c, c) for c in codes)])
    rows.append([])
    rows.append(["SLO Summary"])
    rows.append(["SLO Code", "SLO Description", "Items Touching SLO", "Points Touching SLO"])
    for s in course.get("slos", []):
        hits = [r for r in iter_items(data) if s["code"] in r["item"].get("slo", [])]
        rows.append([s["code"], s["description"], len(hits),
                     sum(r["item"]["points"] for r in hits)])
    return csv_string(rows)


# ----------------------------------------------------------------------
# PDF builder
# ----------------------------------------------------------------------
def build_pdf(data, path):
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import inch
        from reportlab.lib.colors import HexColor, white
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_LEFT, TA_RIGHT
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                        TableStyle, KeepTogether, HRFlowable)
    except ImportError:
        warn("reportlab is not installed, skipped the PDF. Install it with: pip install reportlab")
        return False

    NAVY = HexColor("#1E3D4C")
    GOLD = HexColor("#B8924A")
    TERRA = HexColor("#A0522D")
    GRAY = HexColor("#565E62")
    LINE = HexColor("#C7CDD0")
    TINT = HexColor("#EDF1F3")
    course = data["course"]
    course_start = data["weeks"][0]["start"]

    def st(name, **kw):
        base = dict(fontName="Helvetica", fontSize=9.5, leading=13, textColor=NAVY, alignment=TA_LEFT)
        base.update(kw)
        return ParagraphStyle(name, **base)

    s_eyebrow = st("eyebrow", fontName="Helvetica-Bold", fontSize=8, textColor=TERRA, leading=11)
    s_h1 = st("h1", fontName="Helvetica-Bold", fontSize=23, leading=27)
    s_sub = st("sub", fontName="Helvetica-Bold", fontSize=12, textColor=TERRA, leading=15)
    s_meta = st("meta", fontName="Helvetica-Oblique", fontSize=9, textColor=GRAY, leading=12)
    s_wknum = st("wknum", fontName="Helvetica-Bold", fontSize=30, textColor=TERRA, leading=30)
    s_wktitle = st("wktitle", fontName="Helvetica-Bold", fontSize=13, leading=16)
    s_wkdate = st("wkdate", fontName="Helvetica-Bold", fontSize=8, textColor=GRAY, leading=12)
    s_session = st("session", fontName="Helvetica-Bold", fontSize=7.5, textColor=TERRA, leading=10)
    s_date = st("date", fontName="Helvetica-Bold", fontSize=8, textColor=NAVY, leading=11)
    s_body = st("body", fontSize=9, leading=12)
    s_grade = st("grade", fontName="Helvetica-Bold", fontSize=9, leading=12)
    s_exam = st("exam", fontName="Helvetica-Bold", fontSize=9, textColor=TERRA, leading=12)
    s_label = st("label", fontName="Helvetica-Bold", fontSize=7.5, textColor=TERRA, leading=11)
    s_prep = st("prep", fontSize=8.5, textColor=NAVY, leading=11.5)
    s_link = st("link", fontSize=8, textColor=GRAY, leading=11)

    def two_tone(a, b):
        return ('<font color="%s">%s</font><font color="%s">%s</font>'
                % (NAVY_HEX, a, TERRA_HEX, b))

    story = []
    story.append(Paragraph("%s &nbsp;&middot;&nbsp; %s &nbsp;&middot;&nbsp; %s"
                           % (course["code"], course["title"], course["institution"]), s_eyebrow))
    story.append(Spacer(1, 4))
    story.append(Paragraph(two_tone("Course ", "Schedule"), s_h1))
    story.append(Spacer(1, 2))
    story.append(Paragraph("%s &nbsp;&middot;&nbsp; %s" % (course["title"], course["term"]), s_sub))
    story.append(Spacer(1, 4))
    story.append(Paragraph("%s. Meets %s." % (course["format"], course["meetingPattern"]), s_meta))
    if course.get("isSampleData"):
        story.append(Spacer(1, 6))
        flag = Table([[Paragraph("SAMPLE DATA, replace dates and topics with your real course",
                                 st("flag", fontName="Helvetica-Bold", fontSize=8, textColor=NAVY))]],
                     colWidths=[3.5 * inch])
        flag.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), TINT),
                                  ("BOX", (0, 0), (-1, -1), 0.75, NAVY),
                                  ("TOPPADDING", (0, 0), (-1, -1), 4),
                                  ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                                  ("LEFTPADDING", (0, 0), (-1, -1), 8)]))
        story.append(flag)
    story.append(Spacer(1, 10))
    story.append(HRFlowable(width="100%", thickness=1, color=LINE))
    story.append(Spacer(1, 12))

    # ---- Graded items summary ----
    story.append(Paragraph(two_tone("Graded Items ", "and Due Dates"), s_sub))
    story.append(Spacer(1, 6))
    body = [["Due Date", "Item", "Type", "Covers", "Points"]]
    rows = iter_items(data)
    for r in rows:
        it = r["item"]
        body.append([fmt_short(it["due"]), it["title"],
                     ITEM_TYPES.get(it["type"], it["type"]), it.get("covers", ""), str(it["points"])])
    body.append(["", "Course total", "", "", str(course["totalPoints"])])
    tbl = Table(body, colWidths=[1.0 * inch, 2.55 * inch, 1.5 * inch, 0.75 * inch, 0.7 * inch],
                repeatRows=1)
    ts = [("BACKGROUND", (0, 0), (-1, 0), NAVY),
          ("TEXTCOLOR", (0, 0), (-1, 0), white),
          ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
          ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
          ("FONTSIZE", (0, 0), (-1, -1), 8.5),
          ("TEXTCOLOR", (0, 1), (-1, -1), NAVY),
          ("ALIGN", (4, 0), (4, -1), "CENTER"),
          ("ROWBACKGROUNDS", (0, 1), (-1, -2), [white, HexColor("#FAFAF9")]),
          ("BACKGROUND", (0, -1), (-1, -1), TINT),
          ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
          ("LINEBELOW", (0, 0), (-1, -1), 0.5, LINE),
          ("BOX", (0, 0), (-1, -1), 0.75, NAVY),
          ("TOPPADDING", (0, 0), (-1, -1), 5),
          ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
          ("LEFTPADDING", (0, 0), (-1, -1), 7)]
    for i, r in enumerate(rows, start=1):
        if r["item"]["type"] in EXAM_TYPES:
            ts.append(("LINEBEFORE", (0, i), (0, i), 3, GOLD))
            ts.append(("TEXTCOLOR", (0, i), (-1, i), TERRA))
    tbl.setStyle(TableStyle(ts))
    story.append(tbl)
    story.append(Spacer(1, 18))

    def track_lines(flow, label, items):
        if not items:
            return
        flow.append(Paragraph(label, s_label))
        for x in items:
            flow.append(Paragraph("&bull;&nbsp; " + x["label"], s_prep))
            if x.get("url"):
                flow.append(Paragraph("&nbsp;&nbsp;&nbsp; " + x["url"], s_link))

    def day_card(day):
        flow = []
        is_exam = any(i["type"] in EXAM_TYPES for i in day["items"])
        flow.append(Paragraph("%s &nbsp;&middot;&nbsp; %s"
                              % (day["weekday"], fmt_short(day["date"])), s_date))
        if day.get("sessionType"):
            flow.append(Paragraph(day["sessionType"].upper(), s_session))
        track_lines(flow, "LECTURE", day.get("lecture"))
        track_lines(flow, "LAB", day.get("lab"))
        for it in day["items"]:
            style = s_exam if it["type"] in EXAM_TYPES else s_grade
            flow.append(Spacer(1, 2))
            flow.append(Paragraph(
                "DUE %s &nbsp; %s: %s &nbsp; (%d pts)%s"
                % (fmt_short(it["due"]), ITEM_TYPES.get(it["type"], it["type"]), it["title"],
                   it["points"], ("  covers " + it["covers"]) if it.get("covers") else ""), style))
        flow.append(Spacer(1, 3))
        flow.append(Paragraph("DAILY PREP, recommended, not graded", s_label))
        if day.get("prework"):
            flow.append(Paragraph("&bull;&nbsp; Nightly prework: " + day["prework"]
                                  + " (no link, unlocks after the lecture video)", s_prep))
        if day["date"] != course_start:
            flow.append(Paragraph("&bull;&nbsp; Spaced recall: daily review", s_prep))
        if has_new_cards(day):
            flow.append(Paragraph("&bull;&nbsp; Spaced recall: new cards from today's lab", s_prep))
        card = Table([[flow]], colWidths=[6.5 * inch])
        card.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), white),
            ("BOX", (0, 0), (-1, -1), 0.75, LINE),
            ("LINEBEFORE", (0, 0), (0, 0), 4, GOLD if is_exam else NAVY),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("LEFTPADDING", (0, 0), (-1, -1), 9),
            ("RIGHTPADDING", (0, 0), (-1, -1), 9)]))
        return KeepTogether([card, Spacer(1, 6)])

    def weekend_card(we):
        flow = [Paragraph("WEEKEND &nbsp;&middot;&nbsp; " + fmt_range(we["start"], we["end"]), s_date)]
        if we.get("prework"):
            flow.append(Paragraph("WEEKEND PREWORK", s_label))
            for p in we["prework"]:
                flow.append(Paragraph("&bull;&nbsp; " + p, s_prep))
        if we.get("sunday"):
            flow.append(Paragraph("SUNDAY PREWORK", s_label))
            flow.append(Paragraph("&bull;&nbsp; " + we["sunday"], s_prep))
        flow.append(Paragraph("SPACED RECALL", s_label))
        flow.append(Paragraph("&bull;&nbsp; Daily review, every card due today", s_prep))
        card = Table([[flow]], colWidths=[6.5 * inch])
        card.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), HexColor("#FAFAF9")),
            ("BOX", (0, 0), (-1, -1), 0.75, NAVY),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("LEFTPADDING", (0, 0), (-1, -1), 9),
            ("RIGHTPADDING", (0, 0), (-1, -1), 9)]))
        return KeepTogether([card, Spacer(1, 6)])

    # ---- Per week ----
    for w in data["weeks"]:
        head = Table([[Paragraph("%02d" % w["week"], s_wknum),
                       Paragraph(two_tone("Week %d" % w["week"], ": " + w["theme"]), s_wktitle),
                       Paragraph(fmt_range(w["start"], w["end"]), s_wkdate)]],
                     colWidths=[0.7 * inch, 4.6 * inch, 1.2 * inch])
        head.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                                  ("ALIGN", (2, 0), (2, 0), "RIGHT"),
                                  ("LINEABOVE", (0, 0), (-1, 0), 1, LINE),
                                  ("TOPPADDING", (0, 0), (-1, -1), 8),
                                  ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                                  ("LEFTPADDING", (0, 0), (0, 0), 0)]))
        story.append(KeepTogether([head, Spacer(1, 6), day_card(w["days"][0])]))
        for day in w["days"][1:]:
            story.append(day_card(day))
        if w.get("weekend"):
            story.append(weekend_card(w["weekend"]))
        story.append(Spacer(1, 8))

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(GRAY)
        canvas.drawString(0.75 * inch, 0.5 * inch,
                          "%s %s  |  %s  |  Prepared by Dr. Sharilyn Rennie"
                          % (course["code"], course["title"], course["term"]))
        canvas.drawRightString(7.75 * inch, 0.5 * inch, "Page %d" % doc.page)
        canvas.setStrokeColor(LINE)
        canvas.setLineWidth(0.5)
        canvas.line(0.75 * inch, 0.62 * inch, 7.75 * inch, 0.62 * inch)
        canvas.restoreState()

    doc = SimpleDocTemplate(path, pagesize=letter,
                            leftMargin=0.75 * inch, rightMargin=0.75 * inch,
                            topMargin=0.7 * inch, bottomMargin=0.8 * inch,
                            title="Course Schedule, %s %s" % (course["code"], course["title"]),
                            author="Dr. Sharilyn Rennie")
    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    return True


# ----------------------------------------------------------------------
# Orchestration
# ----------------------------------------------------------------------
def main():
    args = sys.argv[1:]
    source = "auto"
    do_pdf = True
    if "--source" in args:
        source = args[args.index("--source") + 1]
    if "--no-pdf" in args:
        do_pdf = False
    if "--help" in args or "-h" in args:
        print(__doc__)
        return

    if source == "json":
        path = JSON
    elif source == "xlsx":
        path = XLSX
    else:
        path = XLSX if os.path.exists(XLSX) else JSON
    if not os.path.exists(path):
        print("ERROR: no source file found. Expected schedule-data.xlsx or schedule-data.json.")
        sys.exit(1)

    print("Course Schedule Engine (schema 3.0)")
    print("Source: %s" % os.path.basename(path))
    data = read_xlsx(path) if path.endswith(".xlsx") else read_json(path)
    data = normalise(data)

    course = data["course"]
    n_items = sum(len(d["items"]) for w in data["weeks"] for d in w["days"])
    n_lab = sum(1 for r in iter_items(data) if r["item"]["type"] == "lab-exam")
    n_tbl = sum(1 for r in iter_items(data) if r["item"]["type"] == "tbl")
    print("Loaded: %s %s, %d weeks, %d graded items (%d lab exams, %d TBLs), %d points."
          % (course["code"], course["title"], len(data["weeks"]), n_items,
             n_lab, n_tbl, course["totalPoints"]))

    write_json(data, JSON)
    print("  wrote schedule-data.json")
    if update_html(data, HTML):
        print("  refreshed course-schedule.html")

    os.makedirs(EXPORT_DIR, exist_ok=True)
    base = course["code"].replace(" ", "-").lower()
    outputs = {
        base + "-schedule.ics": build_ics(data),
        base + "-spaced-repetition.csv": build_anki(data),
        base + "-study-plan.txt": build_study_plan(data),
        base + "-planner-checklist.md": build_checklist(data),
        base + "-slo-grade-map.csv": build_slo_csv(data),
    }
    for name, text in outputs.items():
        with open(os.path.join(EXPORT_DIR, name), "w", encoding="utf-8", newline="") as f:
            f.write(text)
        print("  wrote exports/%s" % name)

    if do_pdf and build_pdf(data, PDF):
        print("  wrote course-schedule.pdf")

    if WARNINGS:
        print("\nChecks flagged %d item(s) to review:" % len(WARNINGS))
        for w in WARNINGS:
            print("  - " + w)
    else:
        print("\nAll checks passed.")
    print("Done.")


if __name__ == "__main__":
    main()
